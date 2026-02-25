# bot ishlatish uchun 
# 1 python -m venv .venv
# 2 source .venv/Scripts/activate


import asyncio
import os
import sqlite3
import time
from datetime import datetime

import pandas as pd
from openpyxl import Workbook, load_workbook

from aiogram import Bot, Dispatcher, F
from aiogram.filters import Command
from aiogram.types import (
    Message, CallbackQuery,
    ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove,
    ChatJoinRequest, InlineKeyboardMarkup, InlineKeyboardButton
)
from aiogram.utils.keyboard import InlineKeyboardBuilder
from aiogram.fsm.state import StatesGroup, State
from aiogram.fsm.context import FSMContext


# ================= CONFIG =================
BOT_TOKEN = "8385012939:AAEpXYooz8b_AUClguGRSq_9jRDMVTBRjOg"
ADMIN_ID = 5995272629

# To'lovdan keyin kiradigan yopiq kanal/supergroup ID ( -100... )
CHANNEL_ID = -1003668365856

# Ro'yxatdan oldin majburiy obuna bo'ladigan kanal
FORCE_CHANNEL = "@extra_edu"
FORCE_CHANNEL_URL = "https://t.me/extra_edu"

# Excel bazasi (ism-familyalar)
EXCEL_FILE = "students.xlsx"         # ustun: full_name
DB_FILE = "db.sqlite3"

# Ro'yxatdan o'tganlar yoziladigan excel
REG_EXCEL = "registrations.xlsx"

PAYMENT_TEXT = (
    "✅ Bazada topildingiz!\n\n"
    "💳 Karta: 5614 6818 7326 9900\n"
    "💰 Summa: 25 000 so'm\n\n"
    "🧾 Chekni (rasm yoki pdf) qilib yuboring."
)

# Admin tasdiqlagan userlar (join request uchun)
APPROVED_USERS: set[int] = set()


# ================= HELPERS =================
def normalize_name(name: str) -> str:
    return " ".join(str(name).strip().lower().split())


def init_db():
    con = sqlite3.connect(DB_FILE)
    cur = con.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS students(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            full_name TEXT UNIQUE
        )
    """)
    con.commit()
    con.close()


def import_excel_to_db() -> int:
    if not os.path.exists(EXCEL_FILE):
        raise FileNotFoundError(f"Excel topilmadi: {EXCEL_FILE}. Faylni bot.py yoniga qo'ying.")

    df = pd.read_excel(EXCEL_FILE)

    cols = [str(c).strip() for c in df.columns.tolist()]
    if "full_name" not in cols:
        raise ValueError(
            "Excelda 'full_name' ustuni topilmadi.\n"
            f"Topilgan ustunlar: {cols}\n\n"
            "Excel 1-qatorda ustun nomi aynan 'full_name' bo'lishi shart."
        )

    con = sqlite3.connect(DB_FILE)
    cur = con.cursor()

    # qayta import (takror bo'lib ketmasin)
    cur.execute("DELETE FROM students")

    count = 0
    for v in df["full_name"].tolist():
        n = normalize_name(v)
        if not n or n == "nan":
            continue
        cur.execute("INSERT OR IGNORE INTO students(full_name) VALUES (?)", (n,))
        count += 1

    con.commit()
    con.close()
    return count


def student_exists_by_name(full_name: str) -> bool:
    n = normalize_name(full_name)
    con = sqlite3.connect(DB_FILE)
    cur = con.cursor()
    cur.execute("SELECT 1 FROM students WHERE full_name=?", (n,))
    ok = cur.fetchone() is not None
    con.close()
    return ok


def phone_kb():
    return ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="📱 Telegram telefonini yuborish", request_contact=True)]],
        resize_keyboard=True,
        one_time_keyboard=True
    )


def subscribe_kb():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📢 Kanalga obuna bo‘lish", url=FORCE_CHANNEL_URL)],
        [InlineKeyboardButton(text="✅ Tekshirish", callback_data="check_sub")]
    ])


def admin_kb(user_id: int):
    kb = InlineKeyboardBuilder()
    kb.button(text="✅ Tasdiqlash", callback_data=f"approve:{user_id}")
    kb.button(text="❌ Rad etish", callback_data=f"reject:{user_id}")
    kb.adjust(2)
    return kb.as_markup()


def ensure_reg_excel():
    if os.path.exists(REG_EXCEL):
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "registrations"
    ws.append([
        "created_at",
        "tg_id",
        "username",
        "full_name",
        "phone_manual",
        "phone_telegram",
        "school",
        "class",
        "check_result",     # FOUND / NOT_FOUND
        "payment_status"    # NONE / PENDING / APPROVED / REJECTED
    ])
    wb.save(REG_EXCEL)


def append_registration(row: dict):
    ensure_reg_excel()
    wb = load_workbook(REG_EXCEL)
    ws = wb.active
    ws.append([
        row.get("created_at", ""),
        row.get("tg_id", ""),
        row.get("username", ""),
        row.get("full_name", ""),
        row.get("phone_manual", ""),
        row.get("phone_telegram", ""),
        row.get("school", ""),
        row.get("class", ""),
        row.get("check_result", ""),
        row.get("payment_status", ""),
    ])
    wb.save(REG_EXCEL)


async def is_subscribed(user_id: int) -> bool:
    """
    Majburiy kanalga obuna bo'lgan-bo'lmaganini tekshiradi.
    Bot FORCE_CHANNEL kanalga admin qilib qo'shilgan bo'lishi kerak.
    """
    try:
        member = await bot.get_chat_member(chat_id=FORCE_CHANNEL, user_id=user_id)
        # status: member/administrator/creator/left/kicked
        return member.status in ("member", "administrator", "creator")
    except Exception as e:
        # Diagnostika uchun terminalga chiqaramiz
        print("❌ is_subscribed error:", repr(e))
        return False


# ================= STATES =================
class Reg(StatesGroup):
    full_name = State()
    phone_manual = State()
    phone_telegram = State()
    school = State()
    class_ = State()
    waiting_receipt = State()


# ================= BOT =================
bot = Bot(BOT_TOKEN)
dp = Dispatcher()


@dp.message(Command("start"))
async def start(m: Message, state: FSMContext):
    await state.clear()

    # 1) Majburiy kanal obunasi
    if not await is_subscribed(m.from_user.id):
        await m.answer(
            "📌 Ro‘yxatdan o‘tish uchun avval kanalga obuna bo‘ling.\n\n"
            f"{FORCE_CHANNEL_URL}\n\n"
            "Obuna bo‘lgach ✅ Tekshirish tugmasini bosing.",
            reply_markup=subscribe_kb()
        )
        return

    # 2) Ro'yxat boshlandi
    await state.set_state(Reg.full_name)
    await m.answer("✅ Obuna tasdiqlandi!\n\n👤 Ism-familiyangizni kiriting:")


@dp.callback_query(F.data == "check_sub")
async def check_sub(c: CallbackQuery, state: FSMContext):
    if await is_subscribed(c.from_user.id):
        await state.clear()
        await state.set_state(Reg.full_name)
        await c.message.answer("✅ Obuna tasdiqlandi!\n\n👤 Ism-familiyangizni kiriting:")
        await c.answer()
    else:
        await c.answer("❌ Hali kanalga obuna bo‘lmadingiz.", show_alert=True)


@dp.message(Reg.full_name, F.text)
async def step_name(m: Message, state: FSMContext):
    await state.update_data(full_name=m.text.strip())
    await state.set_state(Reg.phone_manual)
    await m.answer("📞 Telefon raqamingizni qo‘lda kiriting:\n(masalan: 901234567 yoki 998901234567)")


@dp.message(Reg.phone_manual, F.text)
async def step_phone_manual(m: Message, state: FSMContext):
    await state.update_data(phone_manual=m.text.strip())
    await state.set_state(Reg.phone_telegram)
    await m.answer("📱 Endi Telegram ochilgan telefoningizni tugma orqali yuboring:", reply_markup=phone_kb())


@dp.message(Reg.phone_telegram, F.contact)
async def step_phone_tg(m: Message, state: FSMContext):
    await state.update_data(phone_telegram=m.contact.phone_number)
    await state.set_state(Reg.school)
    await m.answer("🏫 Maktabingizni yozing:", reply_markup=ReplyKeyboardRemove())


@dp.message(Reg.school, F.text)
async def step_school(m: Message, state: FSMContext):
    await state.update_data(school=m.text.strip())
    await state.set_state(Reg.class_)
    await m.answer("📚 Sinfingizni yozing (masalan: 9-A yoki 10):")


@dp.message(Reg.class_, F.text)
async def step_class(m: Message, state: FSMContext):
    await state.update_data(class_=m.text.strip())
    data = await state.get_data()

    full_name = data.get("full_name", "")
    found = student_exists_by_name(full_name)

    # ✅ Registrations Excelga yozish (topilsa ham, topilmasa ham)
    append_registration({
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "tg_id": m.from_user.id,
        "username": m.from_user.username or "",
        "full_name": full_name,
        "phone_manual": data.get("phone_manual", ""),
        "phone_telegram": data.get("phone_telegram", ""),
        "school": data.get("school", ""),
        "class": data.get("class_", ""),
        "check_result": "FOUND" if found else "NOT_FOUND",
        "payment_status": "PENDING" if found else "NONE",
    })

    if not found:
        await m.answer("❌ Bazadan topilmadingiz (faqat ism-familiya tekshirildi).\nQayta urinish: /start")
        await state.clear()
        return

    await state.set_state(Reg.waiting_receipt)
    await m.answer(PAYMENT_TEXT)


@dp.message(Reg.waiting_receipt, F.photo | F.document)
async def receipt(m: Message, state: FSMContext):
    data = await state.get_data()
    user_id = m.from_user.id

    caption = (
        "🧾 Yangi chek!\n"
        f"👤 Ism: {data.get('full_name','')}\n"
        f"📞 Telefon (qo'lda): {data.get('phone_manual','')}\n"
        f"📱 Telegram telefoni: {data.get('phone_telegram','')}\n"
        f"🏫 Maktab: {data.get('school','')}\n"
        f"📚 Sinf: {data.get('class_','')}\n"
        f"🆔 TG: {user_id} (@{m.from_user.username or 'no_username'})"
    )

    if m.photo:
        await bot.send_photo(ADMIN_ID, m.photo[-1].file_id, caption=caption, reply_markup=admin_kb(user_id))
    else:
        await bot.send_document(ADMIN_ID, m.document.file_id, caption=caption, reply_markup=admin_kb(user_id))

    await m.answer("✅ Chekingiz adminga yuborildi. Tasdiqlansa kanalga kirish chiqadi.")


@dp.callback_query(F.data.startswith("approve:"))
async def approve(c: CallbackQuery):
    if c.from_user.id != ADMIN_ID:
        await c.answer("Siz admin emassiz.", show_alert=True)
        return

    user_id = int(c.data.split(":")[1])
    APPROVED_USERS.add(user_id)

    # ✅ Join request link: forward bo'lsa ham begona kira olmaydi
    invite = await bot.create_chat_invite_link(
        chat_id=CHANNEL_ID,
        creates_join_request=True,
        expire_date=int(time.time()) + 3600  # 1 soat
    )

    await bot.send_message(
        user_id,
        "✅ To‘lov tasdiqlandi!\n"
        "🔗 Kanalga kirish uchun link (so'rov yuboradi):\n"
        f"{invite.invite_link}"
    )
    await c.answer("Tasdiqlandi")


@dp.callback_query(F.data.startswith("reject:"))
async def reject(c: CallbackQuery):
    if c.from_user.id != ADMIN_ID:
        await c.answer("Siz admin emassiz.", show_alert=True)
        return

    user_id = int(c.data.split(":")[1])
    await bot.send_message(user_id, "❌ Chek rad etildi. Iltimos, chekni qayta yuboring.")
    await c.answer("Rad etildi")


@dp.chat_join_request()
async def on_join_request(req: ChatJoinRequest):
    # Faqat admin tasdiqlagan userlarni qabul qilamiz
    if req.from_user.id in APPROVED_USERS:
        await bot.approve_chat_join_request(chat_id=req.chat.id, user_id=req.from_user.id)
        APPROVED_USERS.discard(req.from_user.id)
    else:
        await bot.decline_chat_join_request(chat_id=req.chat.id, user_id=req.from_user.id)


async def main():
    ensure_reg_excel()
    init_db()
    imported = import_excel_to_db()
    print(f"✅ Excel import: {imported} ta ism bazaga yuklandi.")

    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
